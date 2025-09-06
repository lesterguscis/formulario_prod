from flask import Flask, render_template, request, redirect
import os
import base64
import io
from datetime import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image as XLImage
from PIL import Image

app = Flask(__name__)

# Carpeta donde guardaremos los envíos
os.makedirs("submissions", exist_ok=True)

# Nombre del archivo Excel donde guardamos todo
EXCEL_FILE = "submissions/registros.xlsx"

@app.route("/")
def index():
    return render_template("index.html")

@app.route("/submit", methods=["POST"])
def submit():
    # 1. Recibir los datos del formulario
    nombre = request.form["nombre"]
    ci = request.form["ci"]
    fecha_envio = request.form["fecha_envio"]
    provincia = request.form["provincia"]
    correo = request.form["correo"]
    telefono = request.form["telefono"]
    firma_b64 = request.form["firma"]

    # 2. Procesar la firma (Base64 → imagen PNG)
    if firma_b64.startswith("data:image/png;base64,"):
        firma_b64 = firma_b64.replace("data:image/png;base64,", "")
    firma_data = base64.b64decode(firma_b64)

    # Guardamos la firma como archivo PNG
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    firma_filename = f"submissions/firma_{ci}_{timestamp}.png"
    with open(firma_filename, "wb") as f:
        f.write(firma_data)

    # 3. Crear o abrir el archivo Excel
    if not os.path.exists(EXCEL_FILE):
        wb = Workbook()
        ws = wb.active
        ws.title = "Registros"
        # Encabezados
        ws.append(["Nombre y Apellidos", "CI", "Fecha de envío", "Provincia", "Correo", "Teléfono", "Firma"])
        wb.save(EXCEL_FILE)

    wb = load_workbook(EXCEL_FILE)
    ws = wb.active

    # 4. Insertar datos en la siguiente fila
    next_row = ws.max_row + 1
    ws.cell(row=next_row, column=1, value=nombre)
    ws.cell(row=next_row, column=2, value=ci)
    ws.cell(row=next_row, column=3, value=fecha_envio)
    ws.cell(row=next_row, column=4, value=provincia)
    ws.cell(row=next_row, column=5, value=correo)
    ws.cell(row=next_row, column=6, value=telefono)

    # 5. Insertar la firma como imagen en la columna 7
    img = Image.open(io.BytesIO(firma_data))
    img.thumbnail((120, 50))  # Reducir tamaño de la firma
    temp_img = f"submissions/temp_{ci}.png"
    img.save(temp_img)

    xl_img = XLImage(temp_img)
    ws.add_image(xl_img, f"G{next_row}")

    # 6. Guardar Excel
    wb.save(EXCEL_FILE)

    # Borramos temporal
    os.remove(temp_img)

    return "✅ Registro guardado con éxito en Excel."

if __name__ == "__main__":
    app.run(host="0.0.0.0", port=5000, debug=True, use_reloader=False)